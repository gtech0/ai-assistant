from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class ReportService:
    @staticmethod
    def generate_excel_report(project, risks):
        """Генерирует Excel-отчет, полностью учитывающий новую классификацию рисков."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Risks"

        headers = [
            "№", "Название", "Тип", "Фактор", "Категория",
            "Вероятность", "Влияние", "Крит.", "Владелец", "Стратегия"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, risk in enumerate(risks, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=risk.name)
            ws.cell(row=row_idx, column=3, value=risk.risk_type)
            ws.cell(row=row_idx, column=4, value=risk.factor_category)
            ws.cell(row=row_idx, column=5, value=risk.category)
            ws.cell(row=row_idx, column=6, value=risk.probability)
            ws.cell(row=row_idx, column=7, value=risk.impact)
            ws.cell(row=row_idx, column=8, value=risk.criticality)
            ws.cell(row=row_idx, column=9, value=risk.owner)
            ws.cell(row=row_idx, column=10, value=risk.strategy)

        for col in ws.columns:
            max_len = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_len + 2, 40)
            ws.column_dimensions[column].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
